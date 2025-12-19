import os
import openpyxl
from typing import Dict, List, Any, Tuple
import warnings
from pprint import pprint
import re

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')


class CLSData:
    """Class to read and parse CLS/XLSM files for PSSE model building.
    
    Refactored to match logic from Data.py including detailed UG topology and
    all data fields.
    """
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.wb = openpyxl.load_workbook(file_path, data_only=True)
        
        # State for bus renumbering (matching Data.py)
        self.bus_number = 101
        self.JB = 2001
        
        # Sheet counts
        self.sheet_count = {
            'gen': 0,           # 1 General_X
            'xfmr': 0,          # 2 XFMR Impedance_X
            'ohline': 0,        # 3 OH line impedance
            'ug': 0,            # 4 UG collection sys impedance_X
            'gen_imp': 0,       # 5 Generator Impedance_X
            'add_model': 0      # 6 Add Model Setup
        }
        
        # Data containers
        self.gen_data: Dict[str, Dict] = {}
        self.xfmr_2w_data: Dict[str, Dict] = {}
        self.xfmr_3w_data: Dict[str, Dict] = {}
        self.ohline_data: Dict = {}
        self.ug_data: Dict[str, Dict] = {}
        self.gen_imp_data: Dict[str, Dict] = {}
        self.add_model_data: Dict = {}
        self.substation_map: Dict = {} # To track substations like Data.py does
        
        self._count_sheets()
        
    def _count_sheets(self):
        """Count sheets by type."""
        for sheet_name in self.wb.sheetnames:
            if sheet_name.startswith('1 General'):
                self.sheet_count['gen'] += 1
            elif sheet_name.startswith('2 XFMR Impedance'):
                self.sheet_count['xfmr'] += 1
            elif sheet_name.startswith('3 OH line'):
                self.sheet_count['ohline'] += 1
            elif sheet_name.startswith('4 UG collection'):
                self.sheet_count['ug'] += 1
            elif sheet_name.startswith('5 Generator Impedance'):
                self.sheet_count['gen_imp'] += 1
            elif sheet_name.startswith('6 Add Model Setup'):
                self.sheet_count['add_model'] = 1

    def _extract_sheet_index(self, sheet_name: str) -> str:
        """Extract the index suffix from sheet name, handling _ or space separators.
        Matches Data.py behavior partially but more robustly supports 1-9.
        Data.py uses sheet[-1] which is very brittle.
        Here we try to split by _ or space and take last part.
        """
        # Try finding last digit sequence
        match = re.search(r'(\d+)$', sheet_name)
        if match:
            return match.group(1)
        # Fallback to last char if no digit found (matching Data.py worst case?)
        return sheet_name[-1]

    def load_gen_data(self, sheet_name: str) -> Dict:
        """Load data from 1 General_X sheet."""
        sheet = self.wb[sheet_name]
        name = self._extract_sheet_index(sheet_name)
        
        data = {
            'MBASE': float(sheet['B8'].value),
            'LOW_kV': float(sheet['B10'].value),
            'MV_kV': float(sheet['B11'].value),
            'HIGH_kV': float(sheet['B12'].value)
        }
        self.gen_data[name] = data
        return data
    
    def load_xfmr_2w_data(self, sheet_name: str) -> Dict:
        """Load GSU (2-winding) transformer data from 2 XFMR Impedance_X sheet."""
        sheet = self.wb[sheet_name]
        name = self._extract_sheet_index(sheet_name)
        
        data = {
            'R1': round(float(sheet['B12'].value), 6),
            'X1': round(float(sheet['B13'].value), 6),
            'R0': round(float(sheet['B34'].value), 6),
            'X0': round(float(sheet['B35'].value), 6),
            'Vector': str(sheet['B3'].value),
            'MBASE': float(sheet['B14'].value),
            'LOW_kV': float(sheet['C5'].value),
            'MV_kV': float(sheet['B5'].value),
            'NLL': float(sheet['S2'].value),
            'Iext': float(sheet['S3'].value),
            'Tap': float(sheet['S4'].value),
            'Rmax': float(sheet['S5'].value),
            'Rmin': float(sheet['S6'].value),
            'Code': int(sheet['S8'].value) if sheet['S8'].value else 0
        }
        self.xfmr_2w_data[name] = data
        return data
    
    def load_xfmr_3w_data(self, sheet_name: str) -> Dict:
        """Load MPT (3-winding) transformer data from 2 XFMR Impedance_X sheet."""
        sheet = self.wb[sheet_name]
        name = self._extract_sheet_index(sheet_name)
        
        s_val = sheet['T7'].value
        rate = int(s_val.split('/')[-1]) if s_val else 0

        data = {
            'R12': round(float(sheet['D12'].value), 6),
            'X12': round(float(sheet['D13'].value), 6),
            'R23': round(float(sheet['H12'].value), 6),
            'X23': round(float(sheet['H13'].value), 6),
            'R31': round(float(sheet['F12'].value), 6),
            'X31': round(float(sheet['F13'].value), 6),
            'MBASE': float(sheet['D14'].value),
            'R01': round(float(sheet['D34'].value), 6),
            'X01': round(float(sheet['D35'].value), 6),
            'R02': round(float(sheet['H34'].value), 6),
            'X02': round(float(sheet['H35'].value), 6),
            'R03': round(float(sheet['F34'].value), 6),
            'X03': round(float(sheet['F35'].value), 6),
            'Vector': str(sheet['D3'].value),
            'Primary_volt': float(sheet['D5'].value),
            'Secondary_volt': float(sheet['H5'].value),
            'Tertiary_volt': float(sheet['I5'].value),
            'Code': int(sheet['T8'].value) if sheet['T8'].value else 0,
            'NLL': float(sheet['T2'].value),
            'Iext': float(sheet['T3'].value),
            'Tap': float(sheet['T4'].value),
            'Rmax': float(sheet['T5'].value),
            'Rmin': float(sheet['T6'].value),
            'Rate': rate
        }
        self.xfmr_3w_data[name] = data
        return data
    
    def load_ohline_data(self, sheet_name: str) -> Dict:
        """Load overhead line data from 3 OH line impedance sheet."""
        sheet = self.wb[sheet_name]
        data = {
            'LINE': round(float(sheet['G2'].value), 6),
            'kV': float(sheet['E2'].value),
            'RATE': round(float(sheet['I2'].value), 6),
            'R1': round(float(sheet['AD2'].value), 6),
            'X1': round(float(sheet['AE2'].value), 6),
            'R0': round(float(sheet['AF2'].value), 6),
            'X0': round(float(sheet['AG2'].value), 6),
            'B1': round(float(sheet['AJ2'].value), 6),
            'B0': round(float(sheet['AK2'].value), 6)
        }
        self.ohline_data = data
        return data
    
    def load_ug_data(self, sheet_name: str) -> Dict:
        """Load UG collection system data matching Data.py logic."""
        sheet = self.wb[sheet_name]
        name = self._extract_sheet_index(sheet_name)
        
        # Helper vars for topology
        check_bus_list = {}
        gen_type_map = {}
        
        row = 2
        form = {
            'Branch_GSU': {},
            'Gen_Type': {},
            'JB': [],
            'Line': {}, 
            'Rate': {}, 
            'R1': {}, 'X1': {}, 
            'R0': {}, 'X0': {}, 
            'B1': {}, 'B0': {}
        }
        rate_eqv = 0
        
        # 1. Parse Topology
        while True:
            # Using cell syntax to be safe or string formatting
            gen1 = sheet[f'B{row}'].value
            gen2 = sheet[f'C{row}'].value
            
            if gen1 is None or gen2 is None:
                break
                
            line = round(float(sheet[f'G{row}'].value), 6)
            rate = round(float(sheet[f'I{row}'].value), 6)
            r1 = round(float(sheet[f'X{row}'].value), 6)
            x1 = round(float(sheet[f'Y{row}'].value), 6)
            r0 = round(float(sheet[f'Z{row}'].value), 6)
            x0 = round(float(sheet[f'AA{row}'].value), 6)
            b1 = round(float(sheet[f'AB{row}'].value), 6)
            b0 = b1
            type_gen = sheet[f'AN{row}'].value
            
            # --- Bus Renumbering Logic from Data.py ---
            
            # Gen 1 Processing
            if gen1 not in check_bus_list and gen1 % 1000 != 0:
                check_bus_list[gen1] = 1000 + self.bus_number
                form['Branch_GSU'][self.bus_number] = 1000 + self.bus_number
                self.bus_number += 1
            
            # Gen 2 Processing
            if gen2 not in check_bus_list and gen2 % 1000 != 0:
                check_bus_list[gen2] = 1000 + self.bus_number
                form['Branch_GSU'][self.bus_number] = 1000 + self.bus_number
                self.bus_number += 1
                
            # Store Gen Type
            if gen1 not in gen_type_map and gen1 % 1000 != 0:
                gen_type_map[gen1] = type_gen
                
            # JB Logic
            if gen1 % 1000 == 0 and gen1 % 10000 != 0 and gen1 not in check_bus_list:
                check_bus_list[gen1] = self.JB
                form['JB'].append(self.JB)
                self.JB += 1
            
            if gen2 % 1000 == 0 and gen2 % 10000 != 0 and gen2 not in check_bus_list:
                check_bus_list[gen2] = self.JB
                form['JB'].append(self.JB)
                self.JB += 1
                
            # Substation Logic
            if gen1 % 10000 == 0 and gen1 not in self.substation_map:
                self.substation_map[gen1] = name
                check_bus_list[gen1] = gen1
            elif gen1 % 10000 == 0 and gen1 in self.substation_map:
                check_bus_list[gen1] = gen1
                
            if gen2 % 10000 == 0 and gen2 not in self.substation_map:
                self.substation_map[gen2] = name
                check_bus_list[gen2] = gen2
            elif gen2 % 10000 == 0 and gen2 in self.substation_map:
                check_bus_list[gen2] = gen2
            
            # Calc Rate for Equivalents
            if gen2 % 10000 == 0:
                rate_eqv += round(float(sheet[f'I{row}'].value), 6)
                
            # Store Branch Data
            bus_pair = (check_bus_list[gen1], check_bus_list[gen2])
            form['Line'][bus_pair] = line
            form['Rate'][bus_pair] = rate
            form['R1'][bus_pair] = r1
            form['X1'][bus_pair] = x1
            form['R0'][bus_pair] = r0
            form['X0'][bus_pair] = x0
            form['B1'][bus_pair] = b1
            form['B0'][bus_pair] = b0
            
            row += 1
            
        form['Detail'] = form.copy()
        
        # Populate Gen_Type using renumbered IDs
        form['Gen_Type'] = {
            bus_id: gen_type_map.get(gen_id) 
            for gen_id, bus_id in check_bus_list.items() 
            if gen_id in gen_type_map
        }

        # 2. Parse Equivalents
        form_eq = {}
        for i in range(1, sheet.max_row + 1):
            if sheet[f'AH{i}'].value == 'Equivalent R (pu)':
                # Look for the row with data
                for j in range(i + 1, sheet.max_row + 1):
                    if sheet[f'AF{j}'].value is None:
                        continue
                    
                    Name_MPT = sheet[f'AF{j}'].value
                    try:
                        num = int(str(Name_MPT).split()[-1])
                    except:
                        num = 0 # Fallback
                        
                    form_eq = {
                        'R1': sheet[f'AH{j}'].value,
                        'X1': sheet[f'AI{j}'].value,
                        'B1': sheet[f'AJ{j}'].value,
                        'R0': sheet[f'AM{j}'].value,
                        'X0': sheet[f'AN{j}'].value,
                        'B0': sheet[f'AO{j}'].value,
                        'Name_MPT': str(num),
                        'Rate': rate_eqv
                    }
                    break
            if form_eq: break
        
        if form_eq:
             form_eq['Gen_Type'] = form['Gen_Type']
        
        return {
            'Detail': form,
            'Equivalent': form_eq
        }
    
    def load_gen_imp_data(self, sheet_name: str) -> Dict:
        """Load generator impedance data from 5 Generator Impedance_X sheet."""
        sheet = self.wb[sheet_name]
        name = self._extract_sheet_index(sheet_name)
        
        data = {
            'R1': float(sheet['B13'].value),
            'X1': float(sheet['B14'].value),
            'Xsub': float(sheet['B16'].value),
            'Xtrans': float(sheet['B17'].value),
            'Xsyn': float(sheet['B18'].value),
            'R2': float(sheet['B20'].value),
            'X2': float(sheet['B21'].value),
            'R0': float(sheet['B23'].value),
            'X0': float(sheet['B24'].value),
            'Rg': float(sheet['B26'].value),
            'Xg': float(sheet['B27'].value),
            'Rsource': float(sheet['B29'].value),
            'Xsource': float(sheet['B30'].value)
        }
        self.gen_imp_data[name] = data
        return data

    def load_add_model_setup(self, sheet_name: str) -> Dict:
        """Load '6 Add Model Setup' sheet data."""
        sheet = self.wb[sheet_name]
        
        sections = {
            "Reactive Devices": {},
            "Load Devices": {}
        }
        current_section = None
        skip_next = False

        for r in range(1, sheet.max_row + 1):
            b_value = sheet.cell(row=r, column=2).value
            c_value = sheet.cell(row=r, column=3).value
            
            if skip_next:
                skip_next = False
                continue

            if b_value in sections:
                current_section = b_value
                skip_next = True      
                continue

            if current_section:
                if b_value == "" or b_value is None:
                    current_section = None  
                    continue

                sec_dict = sections[current_section]
                item = None
                
                if current_section == "Load Devices":
                    d_value = sheet.cell(row=r, column=4).value  
                    item = [c_value, d_value]   
                else:
                    # Reactive Devices only column C
                    item = c_value

                if b_value in sec_dict:
                    sec_dict[b_value].append(item)
                else:
                    sec_dict[b_value] = [item]
                    
        self.add_model_data = sections
        return sections
    
    def load_all(self):
        """Load all data from the CLS file."""
        # Clean containers for '4UG' to match Data.py nested structure
        self.ug_data_detail = {}
        self.ug_data_equiv = {}
        
        for sheet in self.wb.sheetnames:
            if sheet.startswith("1 General"):
                self.load_gen_data(sheet)
            elif sheet.startswith("2 XFMR Impedance"):
                self.load_xfmr_2w_data(sheet)
                self.load_xfmr_3w_data(sheet)
            elif sheet.startswith("3 OH line impedance"):
                self.load_ohline_data(sheet)
            elif sheet.startswith("4 UG collection"):
                ug_res = self.load_ug_data(sheet)
                name = self._extract_sheet_index(sheet)
                self.ug_data_detail[name] = ug_res['Detail']
                self.ug_data_equiv[name] = ug_res['Equivalent']
            elif sheet.startswith("5 Generator Impedance"):
                self.load_gen_imp_data(sheet)
            elif sheet.startswith("6 Add Model Setup"):
                self.load_add_model_setup(sheet)
                
        return self
    
    def get_all_data(self) -> Dict:
        """Return all loaded data as a dictionary matching Data.py structure."""
        return {
            '1GEN': self.gen_data,
            '2XFMR': self.xfmr_2w_data,
            '3XFMR': self.xfmr_3w_data,
            'OHLINE': self.ohline_data,
            '4UG': {
                'Detail': self.ug_data_detail,
                'Equivalent': self.ug_data_equiv
            },
            '5GEN_IMP': self.gen_imp_data,
            'Subtation': self.substation_map,
            '6Add_Model_Setup': self.add_model_data
        }


if __name__ == "__main__":
    # Test block
    path = r"C:\Users\QHanh\Desktop\INS_Automation_Platform\Backend\DATAs\CLS\2025-04-23_Apricot_Sun_CLS (1).xlsm"
    try:
        cls = CLSData(path)
        cls.load_all()
        data = cls.get_all_data()
        pprint(data, depth=2)
    except Exception as e:
        print(f"Error: {e}")
